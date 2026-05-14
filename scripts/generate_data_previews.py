#!/usr/bin/env python3
"""Generate a single combined pipeline-stages PNG for the README.

Renders all three pipeline stages side-by-side with arrows between them:
  Remote DOL XLSX  →  Normalized CSV  →  Parquet

Output: docs/images/preview-pipeline-stages.png

Usage:
    python3 scripts/generate_data_previews.py [--fy FY] [--quarter Q] [--rows N]
    python3 scripts/generate_data_previews.py --skip-remote   # offline / no download
"""

import argparse
import json
import pathlib
import subprocess
import sys
import tempfile

import pyarrow.csv as pa_csv
import pyarrow.parquet as pq

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.gridspec as gridspec
except ImportError:
    print("matplotlib is required.  pip install matplotlib", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required.  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = pathlib.Path(__file__).resolve().parents[1]
DOCS_IMAGES = ROOT / "docs" / "images"
OUTPUT_PNG = DOCS_IMAGES / "preview-pipeline-stages.png"

DOL_XLSX_URL = (
    "https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/"
    "LCA_Disclosure_Data_FY{fy}_Q{quarter}.xlsx"
)

# Representative subset of raw DOL XLSX columns to display
XLSX_DISPLAY_COLS = [
    "EMPLOYER_NAME",
    "JOB_TITLE",
    "WORKSITE_CITY",
    "WORKSITE_STATE",
    "WAGE_RATE_OF_PAY_FROM",
    "CASE_STATUS",
]

# Columns to display from the normalized CSV / parquet
NORMALIZED_DISPLAY_COLS = [
    "employer",
    "job_title",
    "work_location",
    "wage",
    "status",
    "year",
    "fiscal_quarter",
]

HEADER_COLOR = "#2c7be5"
ROW_COLORS = ["#eef3fb", "#ffffff"]
HEADER_TEXT_COLOR = "white"
MAX_CELL_LEN = 20  # truncate long values so columns stay narrow


def _truncate(value) -> str:
    text = str(value) if value is not None else "—"
    return text[:MAX_CELL_LEN] + "…" if len(text) > MAX_CELL_LEN else text


def _draw_table(ax, headers: list[str], rows: list[list], title: str) -> None:
    ax.axis("off")
    display_rows = [[_truncate(v) for v in row] for row in rows]
    tbl = ax.table(
        cellText=display_rows,
        colLabels=headers,
        cellLoc="left",
        loc="center",
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(7)
    tbl.auto_set_column_width(list(range(len(headers))))
    tbl.scale(1, 1.4)

    for j in range(len(headers)):
        cell = tbl[0, j]
        cell.set_facecolor(HEADER_COLOR)
        cell.set_text_props(color=HEADER_TEXT_COLOR, fontweight="bold")
        cell.set_edgecolor("#1a5fbc")

    for i in range(1, len(rows) + 1):
        for j in range(len(headers)):
            cell = tbl[i, j]
            cell.set_facecolor(ROW_COLORS[(i - 1) % 2])
            cell.set_edgecolor("#cccccc")

    ax.set_title(title, fontsize=8, fontweight="bold", pad=6, color="#1a1a2e")


def _draw_placeholder(ax, title: str, message: str) -> None:
    ax.axis("off")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.add_patch(plt.Rectangle(
        (0.05, 0.15), 0.9, 0.7,
        facecolor="#f5f5f5", edgecolor="#cccccc", linewidth=1,
        transform=ax.transAxes, clip_on=False,
    ))
    ax.text(
        0.5, 0.5, message,
        ha="center", va="center", fontsize=8, color="#999999",
        style="italic", transform=ax.transAxes,
    )
    ax.set_title(title, fontsize=8, fontweight="bold", pad=6, color="#1a1a2e")


def _draw_arrow(ax) -> None:
    ax.axis("off")
    ax.annotate(
        "",
        xy=(0.85, 0.5), xytext=(0.15, 0.5),
        xycoords="axes fraction", textcoords="axes fraction",
        arrowprops=dict(
            arrowstyle="-|>", color="#2c7be5", lw=2.5, mutation_scale=24,
        ),
    )


def render_combined_pipeline_png(
    stages: list[dict],
    output_path: pathlib.Path,
) -> None:
    """Render three pipeline stages side-by-side with arrows into one PNG.

    Each entry in ``stages`` is a dict with keys:
      - title       (str)
      - headers     (list[str] | None)
      - rows        (list[list] | None)
      - placeholder (str)  — shown when headers/rows is None
    """
    # Layout: [table, arrow, table, arrow, table]
    fig = plt.figure(figsize=(28, 6), facecolor="white")
    gs = gridspec.GridSpec(
        1, 5, figure=fig,
        width_ratios=[5, 1, 5, 1, 5],
        wspace=0.04, left=0.01, right=0.99, top=0.85, bottom=0.04,
    )

    table_cols = [0, 2, 4]
    arrow_cols = [1, 3]

    for col_idx, stage in zip(table_cols, stages):
        ax = fig.add_subplot(gs[0, col_idx])
        if stage["headers"] is not None:
            _draw_table(ax, stage["headers"], stage["rows"], stage["title"])
        else:
            _draw_placeholder(ax, stage["title"], stage["placeholder"])

    for col_idx in arrow_cols:
        _draw_arrow(fig.add_subplot(gs[0, col_idx]))

    fig.suptitle(
        "H-1B LCA Pipeline — Sample Data at Each Stage",
        fontsize=12, fontweight="bold", y=0.97, color="#1a1a2e",
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  Saved: {output_path.relative_to(ROOT)}")


# ---------------------------------------------------------------------------
# Stage helpers
# ---------------------------------------------------------------------------

def fetch_xlsx_preview(fy: int, quarter: int, n_rows: int) -> tuple[list[str], list[list]]:
    """Download the DOL XLSX and return (headers, rows) for the display columns."""
    url = DOL_XLSX_URL.format(fy=fy, quarter=quarter)
    print(f"  Downloading {url} …")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = pathlib.Path(tmp.name)

    result = subprocess.run(
        ["curl", "-L", "--fail", "-s", url, "-o", str(tmp_path)],
        capture_output=True,
    )
    if result.returncode != 0:
        tmp_path.unlink(missing_ok=True)
        raise RuntimeError(f"curl failed for {url}")

    wb = load_workbook(filename=tmp_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    raw_header = [str(v).strip() if v is not None else "" for v in next(rows_iter)]

    header_upper = [h.upper() for h in raw_header]
    col_indices: list[int] = []
    found_headers: list[str] = []
    for col in XLSX_DISPLAY_COLS:
        try:
            idx = header_upper.index(col.upper())
            col_indices.append(idx)
            found_headers.append(col)
        except ValueError:
            pass  # column absent in this FY schema

    data_rows: list[list] = []
    for i, row in enumerate(rows_iter):
        if i >= n_rows:
            break
        data_rows.append([row[idx] if idx < len(row) else None for idx in col_indices])

    wb.close()
    tmp_path.unlink(missing_ok=True)
    return found_headers, data_rows


def read_csv_preview(csv_path: pathlib.Path, n_rows: int) -> tuple[list[str], list[list]]:
    table = pa_csv.read_csv(csv_path)
    table = table.slice(0, n_rows)
    available = [c for c in NORMALIZED_DISPLAY_COLS if c in table.schema.names]
    table = table.select(available)
    headers = list(table.schema.names)
    rows = [[table[col][i].as_py() for col in headers] for i in range(table.num_rows)]
    return headers, rows


def read_parquet_preview(parquet_path: pathlib.Path, n_rows: int) -> tuple[list[str], list[list]]:
    available = [
        c for c in NORMALIZED_DISPLAY_COLS
        if c in pq.read_schema(parquet_path).names
    ]
    table = pq.read_table(parquet_path, columns=available)
    table = table.slice(0, n_rows)
    headers = list(table.schema.names)
    rows = [[table[col][i].as_py() for col in headers] for i in range(table.num_rows)]
    return headers, rows


def load_manifest() -> dict:
    manifest_path = ROOT / "data" / "manifest.json"
    if manifest_path.exists():
        return json.loads(manifest_path.read_text())
    return {}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a single combined pipeline-stages PNG for the README."
    )
    parser.add_argument(
        "--fy", type=int, default=None,
        help="Fiscal year for the remote XLSX preview (default: from manifest or 2026).",
    )
    parser.add_argument(
        "--quarter", type=int, default=None,
        help="Quarter for the remote XLSX preview (default: from manifest or 1).",
    )
    parser.add_argument(
        "--rows", type=int, default=5,
        help="Number of preview rows per stage (default: 5).",
    )
    parser.add_argument(
        "--skip-remote", action="store_true",
        help="Skip the remote XLSX download (useful when offline).",
    )
    args = parser.parse_args()

    manifest = load_manifest()
    fy = args.fy or manifest.get("last_fy") or 2026
    quarter = args.quarter or manifest.get("last_quarter") or 1

    stages: list[dict] = []

    # ── Stage 1: Remote DOL XLSX ──────────────────────────────────────────
    print(f"\n[1/3] Remote XLSX  (FY{fy} Q{quarter})")
    stage1: dict = {
        "title": f"① Remote DOL XLSX  (FY{fy} Q{quarter})",
        "headers": None,
        "rows": None,
        "placeholder": "Run with network access\nto populate this stage.",
    }
    if not args.skip_remote:
        try:
            stage1["headers"], stage1["rows"] = fetch_xlsx_preview(fy, quarter, args.rows)
            print("  OK")
        except Exception as exc:
            print(f"  Warning: {exc}")
    else:
        print("  Skipped (--skip-remote).")
    stages.append(stage1)

    # ── Stage 2: Normalized CSV ───────────────────────────────────────────
    print("\n[2/3] Normalized CSV")
    csv_path = ROOT / "data" / "dol_lca_h1b_combined.csv"
    stage2: dict = {
        "title": "② Local Normalized CSV",
        "headers": None,
        "rows": None,
        "placeholder": "Run  npm run fetch:official-data\nto generate this file.",
    }
    if csv_path.exists():
        stage2["headers"], stage2["rows"] = read_csv_preview(csv_path, args.rows)
        print("  OK")
    else:
        print(f"  Not found: {csv_path.relative_to(ROOT)}")
    stages.append(stage2)

    # ── Stage 3: Parquet ─────────────────────────────────────────────────
    print("\n[3/3] Parquet")
    parquet_path = ROOT / "data" / "parquet" / "dol_lca_h1b_combined.parquet"
    stage3: dict = {
        "title": "③ Local Parquet",
        "headers": None,
        "rows": None,
        "placeholder": "Run  npm run build:parquet\nto generate this file.",
    }
    if parquet_path.exists():
        stage3["headers"], stage3["rows"] = read_parquet_preview(parquet_path, args.rows)
        print("  OK")
    else:
        print(f"  Not found: {parquet_path.relative_to(ROOT)}")
    stages.append(stage3)

    print("\nRendering combined PNG …")
    render_combined_pipeline_png(stages, OUTPUT_PNG)
    print("\nDone.")


if __name__ == "__main__":
    main()
