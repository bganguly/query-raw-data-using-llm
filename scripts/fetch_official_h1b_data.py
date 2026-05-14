#!/usr/bin/env python3
import argparse
import csv
import datetime
import json
import os
import pathlib
import re
import subprocess
import sys
from collections.abc import Iterator
from concurrent.futures import ProcessPoolExecutor
from concurrent.futures import ThreadPoolExecutor

from openpyxl import load_workbook

DOL_LCA_XLSX_URL_TEMPLATE = "https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/LCA_Disclosure_Data_FY{fy}_Q{quarter}.xlsx"


def download_file(url: str, target: pathlib.Path) -> bool:
    """Download url to target. Returns True on success, False if the file does not exist (404)."""
    target.parent.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        [
            "curl",
            "-L",
            "--fail",
            "-s",
            "-w",
            "%{http_code}",
            url,
            "-o",
            str(target),
        ],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        if target.exists():
            target.unlink()
        http_code = result.stdout.strip()
        if http_code in ("404", "403", ""):
            return False
        raise RuntimeError(f"Failed to download {url} (HTTP {http_code})")
    return True


def as_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_number(value):
    text = as_text(value)
    if not text:
        return ""
    normalized = re.sub(r"[$,\s]", "", text)
    try:
        return float(normalized)
    except ValueError:
        return ""


def parse_year(row: dict, fallback_year: int) -> int:
    candidates = [
        row.get("CASE_SUBMITTED"),
        row.get("RECEIVED_DATE"),
        row.get("DECISION_DATE"),
        row.get("BEGIN_DATE"),
        row.get("END_DATE"),
        row.get("CASE_RECEIVED_DATE"),
    ]

    for value in candidates:
        text = as_text(value)
        match = re.search(r"(20\d{2})", text)
        if match:
            return int(match.group(1))

    return fallback_year


def resolve_employer_name(row: dict, status: str) -> str:
    employer = as_text(row.get("EMPLOYER_NAME")
                       or row.get("EMPLOYER_NAME_DECLARED"))

    if employer:
        return employer

    alternate_name = as_text(
        row.get("TRADE_NAME_DBA")
        or row.get("SECONDARY_ENTITY_BUSINESS_NAME")
        or row.get("LAWFIRM_NAME_BUSINESS_NAME")
        or row.get("PREPARER_BUSINESS_NAME")
    )

    if alternate_name:
        return alternate_name

    # Many FY2026 Q1 rows are not adjudicated yet and omit employer in the disclosure extract.
    if not status:
        return "N/A - Employer Not Published"

    return ""


def iter_fiscal_quarters(start_fy: int, start_quarter: int, end_fy: int, end_quarter: int) -> Iterator[tuple[int, int]]:
    fy = start_fy
    quarter = start_quarter

    while (fy < end_fy) or (fy == end_fy and quarter <= end_quarter):
        yield fy, quarter

        quarter += 1
        if quarter > 4:
            fy += 1
            quarter = 1


def convert_dol_xlsx_to_normalized_csv(
    source_xlsx: pathlib.Path,
    output_handle,
    write_header: bool,
    fallback_year: int,
    fallback_quarter: int,
    min_calendar_year: int,
    max_rows: int | None,
) -> int:
    workbook = load_workbook(filename=source_xlsx,
                             read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]

    rows_iter = worksheet.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise RuntimeError("DOL workbook has no header row.")

    headers = [as_text(value) for value in header]

    writer = csv.writer(output_handle)
    if write_header:
        writer.writerow(
            [
                "employer",
                "job_title",
                "country",
                "work_location",
                "wage",
                "status",
                "year",
                "fiscal_year",
                "fiscal_quarter",
            ]
        )

    count = 0

    for row_values in rows_iter:
        if not any(value is not None and as_text(value) for value in row_values):
            continue

        row = {headers[index]: row_values[index]
               for index in range(min(len(headers), len(row_values)))}

        visa_class = as_text(row.get("VISA_CLASS")).upper()
        if visa_class and visa_class not in {"H-1B", "H-1B1", "E-3"}:
            continue

        status = as_text(row.get("CASE_STATUS") or row.get("STATUS"))
        employer = resolve_employer_name(row, status)
        job_title = as_text(row.get("JOB_TITLE") or row.get("SOC_TITLE"))
        country = as_text(row.get("WORKSITE_COUNTRY") or row.get(
            "COUNTRY_OF_CITIZENSHIP") or "Unknown")
        city = as_text(row.get("WORKSITE_CITY"))
        state = as_text(row.get("WORKSITE_STATE"))
        work_location = ", ".join([part for part in [city, state] if part])
        wage = (
            parse_number(row.get("WAGE_RATE_OF_PAY_FROM"))
            or parse_number(row.get("WAGE_RATE_OF_PAY_TO"))
            or parse_number(row.get("PREVAILING_WAGE"))
        )
        year = parse_year(row, fallback_year)

        if year < min_calendar_year:
            continue

        writer.writerow(
            [
                employer,
                job_title,
                country,
                work_location,
                wage,
                status,
                year,
                fallback_year,
                fallback_quarter,
            ]
        )
        count += 1

        if max_rows is not None and count >= max_rows:
            break

    workbook.close()
    return count


def normalize_quarter_to_temp_csv(
    source_xlsx: pathlib.Path,
    output_csv: pathlib.Path,
    fallback_year: int,
    fallback_quarter: int,
    min_calendar_year: int,
) -> int:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", newline="", encoding="utf-8") as output_handle:
        return convert_dol_xlsx_to_normalized_csv(
            source_xlsx=source_xlsx,
            output_handle=output_handle,
            write_header=False,
            fallback_year=fallback_year,
            fallback_quarter=fallback_quarter,
            min_calendar_year=min_calendar_year,
            max_rows=None,
        )


def compute_current_fiscal_quarter() -> tuple[int, int]:
    now = datetime.date.today()
    month = now.month
    year = now.year
    if month >= 10:
        return year + 1, 1
    elif month >= 7:
        return year, 4
    elif month >= 4:
        return year, 3
    else:
        return year, 2


def next_fiscal_quarter(fy: int, quarter: int) -> tuple[int, int]:
    if quarter < 4:
        return fy, quarter + 1
    return fy + 1, 1


def read_manifest(manifest_path: pathlib.Path) -> dict | None:
    if not manifest_path.exists():
        return None
    with manifest_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def write_manifest(manifest_path: pathlib.Path, start_fy: int, start_quarter: int, last_fy: int, last_quarter: int) -> None:
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    data = {
        "start_fy": start_fy,
        "start_quarter": start_quarter,
        "last_fy": last_fy,
        "last_quarter": last_quarter,
        "updated_at": datetime.date.today().isoformat(),
    }
    with manifest_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
        f.write("\n")
    print(f"Manifest updated: {manifest_path}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--max-rows", type=int, default=None)
    parser.add_argument("--manifest", default="data/manifest.json",
                        help="Path to pipeline manifest JSON (tracks last processed quarter).")
    parser.add_argument("--output-csv", default="data/dol_lca_h1b_combined.csv",
                        help="Path for the combined normalized CSV output.")
    parser.add_argument("--start-fy", type=int, default=None,
                        help="Override start fiscal year (default: read from manifest or 2020).")
    parser.add_argument("--start-quarter", type=int, default=None,
                        help="Override start fiscal quarter (default: read from manifest or 1).")
    parser.add_argument("--end-fy", type=int, default=None,
                        help="Override end fiscal year (default: auto-computed from current date).")
    parser.add_argument("--end-quarter", type=int, default=None,
                        help="Override end fiscal quarter (default: auto-computed from current date).")
    parser.add_argument("--parallel-downloads", type=int, default=4)
    # Default 2 workers gives a strong speedup while keeping RAM/thermals manageable on older 16 GB Macs.
    parser.add_argument("--parallel-normalize", type=int, default=2)
    parser.add_argument("--min-calendar-year", type=int, default=2020)
    args = parser.parse_args()

    root = pathlib.Path(__file__).resolve().parents[1]
    manifest_path = root / args.manifest
    manifest = read_manifest(manifest_path)

    # Determine effective start quarter.
    explicit_start = args.start_fy is not None and args.start_quarter is not None
    if explicit_start:
        effective_start_fy = args.start_fy
        effective_start_quarter = args.start_quarter
        is_incremental = False
    elif manifest is not None:
        effective_start_fy, effective_start_quarter = next_fiscal_quarter(
            manifest["last_fy"], manifest["last_quarter"]
        )
        is_incremental = True
        print(
            f"[manifest] Last processed: FY{manifest['last_fy']} Q{manifest['last_quarter']}")
        print(
            f"[manifest] Resuming from: FY{effective_start_fy} Q{effective_start_quarter}")
    else:
        effective_start_fy = 2020
        effective_start_quarter = 1
        is_incremental = False

    # Determine effective end quarter.
    if args.end_fy is not None and args.end_quarter is not None:
        effective_end_fy = args.end_fy
        effective_end_quarter = args.end_quarter
    else:
        effective_end_fy, effective_end_quarter = compute_current_fiscal_quarter()
        print(
            f"[manifest] Auto-detected current quarter: FY{effective_end_fy} Q{effective_end_quarter}")

    # Nothing to do if already up to date.
    if (effective_start_fy, effective_start_quarter) > (effective_end_fy, effective_end_quarter):
        print(
            f"Already up to date through FY{effective_end_fy} Q{effective_end_quarter}. Nothing to download.")
        sys.exit(2)

    if effective_start_quarter < 1 or effective_start_quarter > 4:
        raise ValueError("start quarter must be between 1 and 4")
    if effective_end_quarter < 1 or effective_end_quarter > 4:
        raise ValueError("end quarter must be between 1 and 4")
    if args.parallel_downloads < 1:
        raise ValueError("--parallel-downloads must be >= 1")
    if args.parallel_normalize < 1:
        raise ValueError("--parallel-normalize must be >= 1")
    if args.min_calendar_year < 1900:
        raise ValueError("--min-calendar-year must be >= 1900")

    if args.max_rows is not None and args.parallel_normalize > 1:
        print("--max-rows is set; forcing sequential normalization for deterministic truncation.")
        args.parallel_normalize = 1

    data_dir = root / "data"
    source_dir = data_dir / "sources"
    source_dir.mkdir(parents=True, exist_ok=True)

    dol_csv_path = root / args.output_csv

    fiscal_quarters = list(
        iter_fiscal_quarters(effective_start_fy, effective_start_quarter,
                             effective_end_fy, effective_end_quarter)
    )

    quarter_jobs: list[tuple[int, int, pathlib.Path, str]] = []
    for fy, quarter in fiscal_quarters:
        filename = f"LCA_Disclosure_Data_FY{fy}_Q{quarter}.xlsx"
        quarter_url = DOL_LCA_XLSX_URL_TEMPLATE.format(fy=fy, quarter=quarter)
        quarter_xlsx = source_dir / filename
        quarter_jobs.append((fy, quarter, quarter_xlsx, quarter_url))

    print(
        f"Downloading {len(quarter_jobs)} DOL quarter files with parallel downloads={args.parallel_downloads}...")
    with ThreadPoolExecutor(max_workers=args.parallel_downloads) as executor:
        futures = {
            executor.submit(download_file, quarter_url, quarter_xlsx): (fy, quarter, quarter_xlsx, quarter_url)
            for fy, quarter, quarter_xlsx, quarter_url in quarter_jobs
        }
        available_jobs: list[tuple[int, int, pathlib.Path, str]] = []
        for future, job in futures.items():
            fy, quarter, quarter_xlsx, quarter_url = job
            if future.result():
                available_jobs.append(job)
            else:
                print(
                    f"[skip] FY{fy} Q{quarter} not yet published on DOL ({quarter_url})")

    if not available_jobs:
        print("No new DOL quarterly files are available yet. Already up to date.")
        sys.exit(2)

    quarter_jobs = available_jobs

    temp_jobs: list[tuple[int, int, pathlib.Path, pathlib.Path]] = []
    for fy, quarter, quarter_xlsx, _quarter_url in quarter_jobs:
        temp_csv = source_dir / f"normalized_FY{fy}_Q{quarter}.csv"
        temp_jobs.append((fy, quarter, quarter_xlsx, temp_csv))

    if args.parallel_normalize > 1:
        cpu_count = os.cpu_count() or 1
        max_workers = min(args.parallel_normalize, max(
            cpu_count - 1, 1), len(temp_jobs))
        print(
            f"Normalizing {len(temp_jobs)} quarter files in parallel with workers={max_workers} (CPU count={cpu_count})..."
        )
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            futures = [
                executor.submit(
                    normalize_quarter_to_temp_csv,
                    quarter_xlsx,
                    temp_csv,
                    fy,
                    quarter,
                    args.min_calendar_year,
                )
                for fy, quarter, quarter_xlsx, temp_csv in temp_jobs
            ]
            for future in futures:
                future.result()
    else:
        for fy, quarter, quarter_xlsx, temp_csv in temp_jobs:
            print(f"Converting FY{fy} Q{quarter} to normalized rows...")
            normalize_quarter_to_temp_csv(
                source_xlsx=quarter_xlsx,
                output_csv=temp_csv,
                fallback_year=fy,
                fallback_quarter=quarter,
                min_calendar_year=args.min_calendar_year,
            )

    total_rows = 0
    output_written = False
    dol_csv_path.parent.mkdir(parents=True, exist_ok=True)

    # Incremental mode: append new quarter rows to the existing combined CSV.
    csv_mode = "a" if (is_incremental and dol_csv_path.exists()) else "w"
    write_header = csv_mode == "w"

    with dol_csv_path.open(csv_mode, newline="", encoding="utf-8") as output_handle:
        writer = csv.writer(output_handle)
        if write_header:
            writer.writerow(
                [
                    "employer",
                    "job_title",
                    "country",
                    "work_location",
                    "wage",
                    "status",
                    "year",
                    "fiscal_year",
                    "fiscal_quarter",
                ]
            )

        for _fy, _quarter, quarter_xlsx, temp_csv in temp_jobs:
            with temp_csv.open("r", newline="", encoding="utf-8") as temp_handle:
                reader = csv.reader(temp_handle)
                for row in reader:
                    if args.max_rows is not None and total_rows >= args.max_rows:
                        break
                    writer.writerow(row)
                    total_rows += 1

            output_written = output_written or total_rows > 0

            if temp_csv.exists():
                temp_csv.unlink()
            if quarter_xlsx.exists():
                quarter_xlsx.unlink()

            if args.max_rows is not None and total_rows >= args.max_rows:
                break

    if not output_written:
        raise RuntimeError(
            "No DOL rows were written. Check source files and conversion logic.")

    # Remove older dated-name CSVs to avoid confusion.
    legacy_csv = data_dir / "dol_lca_h1b_fy2026_q1.csv"
    if legacy_csv.exists() and legacy_csv != dol_csv_path:
        legacy_csv.unlink()

    if source_dir.exists() and not any(source_dir.iterdir()):
        source_dir.rmdir()

    # Update manifest with the new last-processed quarter.
    manifest_start_fy = manifest["start_fy"] if manifest else effective_start_fy
    manifest_start_quarter = manifest["start_quarter"] if manifest else effective_start_quarter
    write_manifest(manifest_path, manifest_start_fy,
                   manifest_start_quarter, effective_end_fy, effective_end_quarter)

    print("Done.")
    print(
        f"DOL fiscal quarter range processed this run: FY{effective_start_fy} Q{effective_start_quarter} -> FY{effective_end_fy} Q{effective_end_quarter}")
    print(f"Minimum included calendar year: {args.min_calendar_year}")
    print(f"DOL normalized CSV rows (this run): {total_rows}")
    print(f"DOL combined CSV: {dol_csv_path}")


if __name__ == "__main__":
    main()
